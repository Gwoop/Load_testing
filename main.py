import random
import threading
from tkinter import *
from tkinter import ttk
from functools import partial
import psutil#1
#import pymysql
import openpyxl#2
import subprocess
from threading import Thread
from threading import Timer
import datetime
from datetime import date
from multiprocessing import Process
import os
import cx_Oracle
import time
from tkinter.messagebox import showinfo,showerror,showwarning
import schedule
import wx #4

decryptedAccses = []

class TimeEditor(Tk):
    def __init__(self):
        super().__init__()
        self.title("Таймер запуска")
        self.geometry("450x180")

        self.btn1 = ttk.Button(self, text="Запуск")
        self.btn1["command"] = self.start
        self.btn1.grid(column=0, row=0, pady=20, padx=35)

        self.btn2 = ttk.Button(self, text="Остановка")
        self.btn2["command"] = self.end
        self.btn2.grid(column=1, row=0, pady=20, padx=35)

        self.btn2 = ttk.Button(self, text="Назад")
        self.btn2["command"] = self.back
        self.btn2.grid(column=2, row=0, pady=20, padx=35)

        self.labelStatus = Label(self, text="Введите сюда даты запуска", font=("Arial Bold", 12), justify='center')
        self.labelStatus.grid(column=0, columnspan=3, row=1, pady=20, padx=35)

        self.CommandText = Entry(self, width=20, font=("Arial Bold", 12))
        self.CommandText.grid(column=0, columnspan=3, row=2, pady=20, padx=35)



        self.loaddata() #метод загрузки окна

    def start(self):
        # global p
        # p.kill()
        from datetime import timedelta
        from datetime import datetime
        now = datetime.now()
        run_at = now + timedelta(minutes=1)
        delay = (run_at - now).total_seconds()

        threading.Timer(5, threads(int("60"), "self.nameDB.get()")).start()


        pass

    def end(self):
        pass

    def loaddata(self):
        pass

    def back(self):
        self.destroy()
        window = MainWindow()

#Окно для взаимодесвия с конфиг файлом
class EditXLSS(Tk):
    def __init__(self):
        super().__init__()
        # тут я его настраивать начинаю
        self.title("Настройка запросов")
        self.geometry("920x500")

        self.CommandText = Text(width=90,height=20,font=("Arial Bold", 12))
        self.CommandText.grid(column=0, columnspan=4 ,row=1,pady=20,padx=35)

        self.btn1 = ttk.Button(self, text="Обновить данные")
        self.btn1["command"] = self.update
        self.btn1.grid(column=0,row=0,pady=20,padx=35)

        self.btn2 = ttk.Button(self, text="Сохранить данные")
        self.btn2["command"] = self.load
        self.btn2.grid(column=1, row=0, pady=20, padx=35)

        self.btn3 = ttk.Button(self, text="Назад в меню")
        self.btn3["command"] = self.back
        self.btn3.grid(column=2, row=0, pady=20, padx=35)
        self.update()

        self.entryNameDb = ttk.Entry(self, width=20, font=("Arial Bold", 12))  # поле ввода
        self.entryNameDb.grid(column= 3, row=0, columnspan=3, pady=10)
        # через partial передаю нормально аргументы в метод иначе оч плохо всё будет

        self.entryNameDb.insert(0, decryptedAccses[2])
    def back(self):
        self.destroy()
        window = MainWindow()

    def load(self):
        test = self.CommandText.get('1.0', END)
        mas = []
        mas = test.split('\n')
        mas = list(filter(None,mas))


        user = decryptedAccses[1]
        password = decryptedAccses[0]
        dnName = self.entryNameDb.get()
        try:
            connection = cx_Oracle.connect(user + "/" + password + "@" + dnName, encoding="UTF-8")
            print(connection)
            cursor = connection.cursor()
            try:

                for i in mas:
                    if(i != None or i != "" or i != ''):
                        dnss = f"""
                                           {i}"""
                        cursor.execute(str(dnss))
                        ver = cursor.fetchone()

                wb = openpyxl.load_workbook('sql.xlsx')
                sheet = wb.active

                for i in range(len(mas)):
                    sheet["A" + str(i + 1)] = mas[i]
                wb.save('sql.xlsx')
                showinfo(title="Данные сохраненны",
                        message="Данные успешно сохраненны")
            except:
                showerror(title="Ошибка",
                          message="Какая-то из запросов является не рабочим и выдаёт ошибку\nПока запросы не будут корректны, данные не сохранятся\nПроверьте что вы ввели")
        except:
            showerror(title="Ошибка",
                      message="Не удалось подключиться к базе даныых")


    def update(self):
        gg = readxlsx()
        self.CommandText.delete('1.0', END)
        for i in range(len(gg)):
            self.CommandText.insert(INSERT, gg[i] + '\n')



#метод для создания потоков и управления ими
def threads(maxLimit,nameDB):
    zgluchka =0
    xlsx = readxlsx()
    i = 0
    T = 1  # стартовое число потоков
    countermax = T
    sended_request = 0
    while zgluchka == 0:
        #это занимает некоторое время, надо что-то придумать с этим
        #currentCPU = GetCpuPersents()
        #для того чтобы узнать GPU -> psutil.virtual_memory()[2]
        if psutil.virtual_memory()[2] < maxLimit:
            T += 1
        if psutil.virtual_memory()[2] > maxLimit:  # верхний порог нагрузки
            T -= 1
        if T > countermax:  #
            countermax = T
        if T <= 0:
            T = 1
        time = Timer(5, log, args=(T, countermax, sended_request,))
        time.start()

        threads = []
        for n in range(int(T)):
            t = Thread(target=bdoracle, args=(xlsx,nameDB,), daemon=False)
            threads.append(t)
            t.start()
            sended_request = sended_request + 1


#метод для поключения к БД и отправке запросов
def bdoracle(xls, dnName):
    user = decryptedAccses[1]
    password = decryptedAccses[0]
    connection = cx_Oracle.connect(user + "/" + password + "@" + dnName, encoding="UTF-8")
    print(connection)
    cursor = connection.cursor()

    for i in xls:
        dnss = f"""
               {i}"""
        cursor.execute(dnss)
        ver = cursor.fetchone()


def ReadTxtAndBackMassive():
    massive = []
    d =  open("DecrypedData.txt", "r")
    for line in d:
        lineClear = line.replace("\n","")
        massive.append(lineClear)
    d.close()
    f = open('DecrypedData.txt', 'w')
    f.close()
    return massive


#метод для расшифровки данных для доступа к БД
def ReadTxtAndBackMassiveWithout():
    massive = []
    d =  open("DecrypedData.txt", "r")
    for line in d:
        lineClear = line.replace("\n","")
        massive.append(lineClear)
    d.close()
    return massive



#Метод для получения октуальных данных о нагрузке
def GetCpuPersents():
    output = str(subprocess.check_output('wmic cpu get loadpercentage'))
    nowCpu = int(output[24] + output[25])
    return nowCpu


#метод для запуска процесса в независимом пространсве от основной программы
def proc_start(maxLimit, BD):
    p_to_start = Process(target=threads, args=(int(maxLimit), BD,), daemon=False)
    p_to_start.start()
    return p_to_start


#метод для остановки процесса
def proc_stop(p_to_stop):
    p_to_stop.kill()




#главное окно приложения
class MainWindow(Tk):

    #метод создания окна
    def __init__(self):
        super().__init__()
        #главная настройка окна
        self.title("Тестирование")  # заголовок окна
        self.geometry(
            '790x230')
        #self.resizable(height=False, width=False)
        self.label1 = Label(self, text="В поле ниже введите максимальную\nнагрузку на ПК в процентах",
                            font=("Arial Bold", 12), justify='left')  # просто информационное поле
        self.label1.grid(column=0, row=0, padx=10)  # это пишется чтобы элемент был хотя-бы виден и настроен по сетке
        self.labelStatus = Label(self, text="Остановлен", font=("Arial Bold", 12), justify='right')
        self.labelStatus.grid(column=2, row=0,padx=10)  # это поле будем изменять. Оно нужно для отображения статуса программы
        self.procents = Entry(self, width=20, font=("Arial Bold", 12))  # поле ввода нагрузки в процентах
        self.procents.grid(column=0, row=1, pady=10, sticky="w", padx=10)
        btnStart = ttk.Button(self, text="Старт")
        btnStart.grid(column=0, row=2, pady=10, sticky="w", padx=10)
        btnStart["command"] = partial(self.btnStart)

        self.labelStatus = Label(self, text="В поле ниже введите название базы данных", font=("Arial Bold", 12), justify='right')
        self.labelStatus.grid(column=1, row=0,padx=10)
        self.nameDB = Entry(self, width=20, font=("Arial Bold", 12))  # поле ввода названия БД
        self.nameDB.grid(column=1, row=1, pady=10, sticky="n", padx=10)

        #вставляю в поле ввода названия БД - название БД взятого из файла
        if(onLocal == False):
            print()
            self.nameDB.insert(0,str( decryptedAccses[2]));

        btnStop = ttk.Button(self, text="Остановить")
        btnStop.grid(column=2, row=2, pady=10, sticky="w")
        btnStop["command"] = self.btnStop

        btnSave = ttk.Button(self, text="Переход в окно настройки запросов")
        btnSave.grid(column=0, row=3, pady=10, sticky="w", padx=10)
        btnSave["command"] = self.gotoZapros

        btnSave = ttk.Button(self, text="Переход в окно настройки таймера")
        btnSave.grid(column=1, row=3, pady=10, sticky="w", padx=10)
        btnSave["command"] = self.gotoTimer




    def ProgramDelay(self):
        global p
        from datetime import datetime
        now = datetime.now()
        from datetime import timedelta
        run_at = now + timedelta(minutes=1)
        delay = (run_at - now).total_seconds()
        proc_stop(p)
        time.sleep(delay)
        p = proc_start(self.procents.get(),self.nameDB.get())



    #старт теста
    def btnStart(self):
        self.labelStatus.config(text="Работает")
        xlsx = readxlsx()

        if(len(self.nameDB.get()) == 0):
            showerror(title="Ошибка",
                      message="Введите название базы данных к которой хотите подключиться")
        else:
            if(len(self.procents.get()) == 0):
                showerror(title="Ошибка",
                          message="Введите число для нагрузки")
            else:
                global p
                p = proc_start(self.procents.get(),self.nameDB.get())
                #p = threads(int(self.procents.get()),self.nameDB.get())





    def gotoZapros(self):
        self.destroy()
        window = EditXLSS()

    #Кнопка для запска через определённое время или для возобновления работы через определённое время
    def gotoTimer(self):
        global p
        from datetime import datetime
        now = datetime.now()
        from datetime import timedelta
        run_at = now + timedelta(minutes=1)
        delay = (run_at - now).total_seconds()
        proc_stop(p)
        time.sleep(delay)
        p = proc_start(self.procents.get(), self.nameDB.get())


    #Кнопка для остановки нагрузки
    def btnStop(self):
        self.labelStatus.config(text="Остановлен")
        global p
        proc_stop(p)







# Метод для считывая файла с запросами sql.xlsx
def readxlsx():
    wb = openpyxl.load_workbook('sql.xlsx')
    sheet = wb.active
    xlsx = []
    for i in range(sheet.max_row):
        xlsx.append(str(sheet["A" + str(i + 1)].value))
    return xlsx

# Метод для создания логов
def log(num,nummax,sended_request):
    current_date = date.today()
    dt_now = datetime.datetime.now()
    f = open(str(current_date)+'.txt','a')
    f.write(str(dt_now) + " active_thread: " + str(num)  + " max_thread: " + str(nummax) + " sended_request: " + str(sended_request) + "\n")
    f.close()


onLocal = True #если True, то запуск будет без проверок на подключение и без дешифровки данных для подключения
#если False, то со всеми ними

# Метод для обработки нажатия кновки, который проверяет наличие всех необходимых компонентов, для запуска программы
def click():
    #чтение дешифрованных данных и их удаление
    if(onLocal == False):
        subprocess.Popen('Project1.exe')
        time.sleep(5)
        global decryptedAccses
        decryptedAccses = ReadTxtAndBackMassive()
        xlsx = readxlsx()
        if (xlsx != [] and decryptedAccses != []):
            root.destroy()
            window = MainWindow()
        else:
            showerror(title="Ошибка",
                      message="К сожалению первичная настройка приложения прошла не успешно\nПроверьте, всё-ли верно вы настроили")
    else:
        root.destroy()
        window = MainWindow()



if __name__ == '__main__':
        root = Tk()
        root.title("Окно запуска")
        root.geometry("250x200")
        open_button = ttk.Button(text="Запуск", command=click)
        open_button.pack(anchor="center", expand=1)
        root.mainloop()
